from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
import openpyxl
import logging
import requests
from datetime import datetime, timedelta
from django.core.cache import cache
from django.utils import timezone
from django.contrib.auth.decorators import login_required

from apps.vacancies.models import Vacancy, GeminiResult, GeminiPrompt, TaskQueue, ExchangeRate, UserProfile
from apps.vacancies.forms import GeminiInputForm, GeminiPromptForm, UserProfileForm, VacancyEditForm
from django.contrib import messages

# Для сводной статистики
from statistics import median
from collections import defaultdict

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q

# Настраиваем логирование
logger = logging.getLogger(__name__)


def vacancy_detail(request, vacancy_id):
    vacancy = get_object_or_404(Vacancy, pk=vacancy_id, is_active=True)
    form = VacancyEditForm(instance=vacancy)
    edit_history = {
        'editor': vacancy.last_edited_by.get_full_name() if vacancy.last_edited_by else None,
        'edited_at': vacancy.last_edited_at
    }
    return render(request, 'vacancies/detail.html', {
        'vacancy': vacancy,
        'form': form,
        'edit_history': edit_history
    })


def index(request):
    search_query = request.GET.get('search', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    per_page = request.GET.get('per_page', '50')  # По умолчанию 50 записей
    
    if search_query:
        vacancies = Vacancy.objects.filter(
            Q(company__icontains=search_query) |
            Q(geo__icontains=search_query) |
            Q(specialization__icontains=search_query) |
            Q(grade__icontains=search_query) |
            Q(description__icontains=search_query),
            is_active=True
        ).order_by('-date_posted')
    else:
        vacancies = Vacancy.objects.filter(is_active=True).order_by('-date_posted')
    
    if date_from:
        df = datetime.strptime(date_from, '%Y-%m-%d')
        vacancies = vacancies.filter(date_posted__gte=df)
    if date_to:
        dt = datetime.strptime(date_to, '%Y-%m-%d')
        vacancies = vacancies.filter(date_posted__lte=dt)
    
    # Настраиваем пагинацию
    if per_page == 'all':
        # Если выбрано "все", не используем пагинацию
        paginated_vacancies = vacancies
    else:
        try:
            # Преобразуем per_page в число
            items_per_page = int(per_page)
            paginator = Paginator(vacancies, items_per_page)
            page = request.GET.get('page')
            
            try:
                paginated_vacancies = paginator.get_page(page)
            except PageNotAnInteger:
                # Если страница не является целым числом, возвращаем первую страницу
                paginated_vacancies = paginator.get_page(1)
            except EmptyPage:
                # Если страница больше максимальной, возвращаем последнюю страницу
                paginated_vacancies = paginator.get_page(paginator.num_pages)
        except ValueError:
            # Если per_page не является числом и не 'all', используем значение по умолчанию
            paginator = Paginator(vacancies, 50)
            paginated_vacancies = paginator.get_page(1)
    
    context = {
        'vacancies': paginated_vacancies,
        'search_query': search_query,
    }
    
    return render(request, 'vacancies/index.html', context)


def add_vacancy(request):
    from apps.vacancies.forms import VacancyForm  # Предполагается, что форма добавления вакансии уже есть
    if request.method == 'POST':
        form = VacancyForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('index')
    else:
        form = VacancyForm()
    return render(request, 'vacancies/add.html', {'form': form})


def export_vacancies(request):
    wb = openpyxl.Workbook()
    ws = wb.active

    # Список полей в нужном порядке.
    fields = [
        'company', 'geo', 'specialization', 'grade',
        'salary_min', 'salary_max', 'bonus', 'bonus_conditions',
        'currency', 'gross_net', 'work_format', 'date_posted',
        'source', 'author'
    ]

    # Формирование заголовков столбцов с использованием verbose_name из модели
    header = [Vacancy._meta.get_field(field).verbose_name for field in fields]
    ws.append(header)

    # Заполнение данных вакансий
    for vac in Vacancy.objects.filter(is_active=True):
        row = [getattr(vac, field) for field in fields]
        ws.append(row)

    response = HttpResponse(content_type="application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename="vacancies.xlsx"'
    wb.save(response)
    return response


def gemini(request):
    """
    Представление для обработки ввода через форму Gemini.
    Используется кастомный промпт из модели GeminiPrompt (если он задан), а иначе значение по умолчанию.
    """
    if request.method == 'POST':
        form = GeminiInputForm(request.POST)
        if form.is_valid():
            TaskQueue.create(data=form.cleaned_data['text'], priority=TaskQueue.Priority.CRITICAL)
            return redirect("gemini_result")
    else:
        form = GeminiInputForm()
    return render(request, "vacancies/gemini_form.html", {"form": form})


def change_prompt(request):
    """
    Представление для изменения промпта для Gemini AI.
    Если промпт уже существует, форма предзаполнена текущим значением.
    После отправки формы промпт обновляется.
    """
    gemini_prompt_obj = GeminiPrompt.objects.first()
    if not gemini_prompt_obj:
        default_prompt = (
            "Проанализируй следующий текст о вакансии и верни результат в виде JSON-объекта с указанными ключами. JSON должен содержать следующие поля:\n"
            "'Company': название компании.\n"
            "'Geo': местоположение компании или вакансии. Определяется по тексту. В ответе возвращается название всех стран, где есть офисы/представление, через запятую.\n"
            "'Specialization': специализация или направление работы. Выбирается из списка наиболее подходящая из названия или описания :Product Owner, Digital Product Owner, Project Manager, Scrum Master, Business Analyst, System Analyst, UX Researcher, UX/UI Designer, UX Writer, Enterprise Architect, Security Architect, Data Architect, SecOps Engineer, AppSec Engineer, DevOps Engineer, System Administrator, Relational Database Administrator, NoSQL Database Administrator, Network Engineer (Cisco), Cyber Security Engineer, Web Developer, Android Developer, iOS Developer, Kotlin Multiplatform Engineer, Java Developer, C# / .NET Developer, Python Developer, Erlang / Elixir Developer, Full-Stack Developer, RPA Developer, HCL/Lotus Notes Developer, 1C Developer, EQ / RPG Analyst, EQ / RPG Developer, Web QA, Mobile QA, Load / Performance QA, Release Manager, Data Analyst, Web Analyst, MLOps Engineer, DS / ML Engineer, AI Engineer, Technical Support Engineer, IT Auditor, Monitoring Engineer, Penetration Tester, Security Auditor, Digital Marketer, Internet Marketer, SOC Engineer, SIEM Engineer, Data Engineer, BI Analyst, Marketing Analyst.\n"
            "'Grade': уровень должности (если не указан, то определяется на основании описания и требуемого опыта).\n"
            "'Salary Min': минимальная зарплата (если не указана, то None).\n"
            "'Salary Max': максимальная зарплата (если не указана, то None).\n"
            "'Bonus': размер бонуса, если нет, то None.\n"
            "'Bonus Conditions': условия предоставления бонуса, если нет, то None.\n"
            "'Currency': валюта расчёта в формате USD, RUB и так далее.\n"
            "'Gross/Net': информация о типе оплаты (до вычета/после вычета налогов, net=на руки, чистыми, gross=грязыми, до вычета налогов. Если не указано, то по определению gross)\n"
            "'Work Format': формат работы (возможные: удаленная, офис, релокация, гибрид).\n"
            "'Date Posted': дата публикации вакансии (сейчас).\n"
            "'Source': источник вакансии (из формы, hh, tg)\n"
            "'Author': автор публикации вакансии.\n\n"
            "Не добавляй никаких дополнительных полей или комментариев. Верни только валидный JSON-объект.\n\n"
            "Текст: "
        )
        gemini_prompt_obj = GeminiPrompt.objects.create(prompt_text=default_prompt)

    if request.method == 'POST':
        form = GeminiPromptForm(request.POST, instance=gemini_prompt_obj)
        if form.is_valid():
            form.save()
            return redirect('gemini')
    else:
        form = GeminiPromptForm(instance=gemini_prompt_obj)
    return render(request, "vacancies/change_prompt.html", {"form": form})


def gemini_result(request):
    results = GeminiResult.objects.order_by('-created_at')
    return render(request, "vacancies/gemini_result.html", {"results": results})


#
# -------------------- ВАЖНО: ВАША СТАРАЯ ФУНКЦИЯ pivot_summary => МЫ ДОПОЛНЯЕМ ЕЁ ВЫВОДОМ BYN --------------------
#

def pivot_summary(request):
    """
    Сводная таблица (две части):
      1) Gross (как раньше)
      2) Переводим в BYN
    """
    # Считываем фильтры из GET-параметров
    spec_filter = request.GET.get('specialization', '').strip()
    grade_filter = request.GET.get('grade', '').strip()
    geo_filter = request.GET.get('geo', '').strip()

    # Формируем начальный queryset с фильтрацией по последним 92 дням и только активные вакансии
    today = timezone.now().date()
    ninety_two_days_ago = today - timedelta(days=91)  # 91 день назад включительно
    qs = Vacancy.objects.filter(date_posted__gte=ninety_two_days_ago, is_active=True)

    # Применяем простую фильтрацию (по точному совпадению)
    if spec_filter:
        qs = qs.filter(specialization__iexact=spec_filter)
    if grade_filter:
        qs = qs.filter(grade__iexact=grade_filter)
    if geo_filter:
        qs = qs.filter(geo__iexact=geo_filter)

    # Собираем нужные поля
    data = qs.values(
        'specialization',
        'grade',
        'geo',
        'currency',
        'salary_min',
        'salary_max',
        'gross_net'
    ).order_by('specialization', 'grade', 'geo', 'currency')

    def to_gross(value, gross_net_flag):
        """
        Если gross_net == net, конвертируем value в gross путём деления на 0.86 и округления.
        Иначе оставляем как есть.
        Если value=None, возвращаем None.
        """
        if value is None:
            return None
        if gross_net_flag and gross_net_flag.lower() == 'net':
            return int(round(value / 0.86))
        return value

    # Функция для конвертации в BYN (пример «заглушки»)
    def convert_to_byn(amount, ccy):
        """
        Конвертирует сумму в BYN по актуальному курсу НБРБ.
        
        Args:
            amount: сумма для конвертации
            ccy: валюта (USD, EUR, RUB, BYN и др.)
        
        Returns:
            float: сумма в BYN или None если конвертация невозможна
        """
        if amount is None or not ccy:
            return None
        
        if ccy == 'BYN':
            return amount
        
        rates = get_exchange_rates()
        
        if ccy in rates:
            return amount * rates[ccy]
        
        # Если валюта неизвестна, логируем это и возвращаем None
        logger.warning(f"Неизвестная валюта для конвертации: {ccy}")
        return None

    # 1) Сгруппируем по (spec, grade, geo, currency) -> получим "gross" значения
    groups_gross = defaultdict(lambda: {
        'sal_min': [],
        'sal_max': [],
        'all_salaries': [],
    })

    for row in data:
        spec = row['specialization']
        grd = row['grade']
        current_geo = row['geo']
        ccy = row['currency']
        gross_net_val = row['gross_net'] or ''

        sal_min_gross = to_gross(row['salary_min'], gross_net_val)
        sal_max_gross = to_gross(row['salary_max'], gross_net_val)

        if sal_min_gross is not None:
            groups_gross[(spec, grd, current_geo, ccy)]['sal_min'].append(sal_min_gross)
            groups_gross[(spec, grd, current_geo, ccy)]['all_salaries'].append(sal_min_gross)
        if sal_max_gross is not None:
            groups_gross[(spec, grd, current_geo, ccy)]['sal_max'].append(sal_max_gross)
            groups_gross[(spec, grd, current_geo, ccy)]['all_salaries'].append(sal_max_gross)

    # 2) Превращаем groups_gross в список results_gross
    results_gross = []
    for (spec, grd, current_geo, ccy), values_dict in groups_gross.items():
        min_list = values_dict['sal_min']
        max_list = values_dict['sal_max']
        all_list = values_dict['all_salaries']

        if min_list:
            min_median = median(min_list)
        else:
            min_median = None

        if max_list:
            max_median = median(max_list)
        else:
            max_median = None

        if all_list:
            market_median = median(all_list)
        else:
            market_median = None

        results_gross.append({
            'specialization': spec,
            'grade': grd,
            'geo': current_geo,
            'currency': ccy,  # оставляем "оригинальную" валюту
            'min_median': int(round(min_median)) if min_median is not None else None,
            'max_median': int(round(max_median)) if max_median is not None else None,
            'market_median': int(round(market_median)) if market_median is not None else None,
        })

    #
    # -- Теперь для BYN -- (делаем ещё один проход, или можно было совместить)
    #

    groups_byn = defaultdict(lambda: {
        'sal_min': [],
        'sal_max': [],
        'all_salaries': [],
    })

    for row in data:
        spec = row['specialization']
        grd = row['grade']
        current_geo = row['geo']
        ccy = row['currency']
        gross_net_val = row['gross_net'] or ''

        sal_min_gross = to_gross(row['salary_min'], gross_net_val)
        sal_max_gross = to_gross(row['salary_max'], gross_net_val)

        # Конвертируем именно в BYN
        sal_min_byn = convert_to_byn(sal_min_gross, ccy) if sal_min_gross is not None else None
        sal_max_byn = convert_to_byn(sal_max_gross, ccy) if sal_max_gross is not None else None

        if sal_min_byn is not None:
            groups_byn[(spec, grd, current_geo)]['sal_min'].append(sal_min_byn)
            groups_byn[(spec, grd, current_geo)]['all_salaries'].append(sal_min_byn)
        if sal_max_byn is not None:
            groups_byn[(spec, grd, current_geo)]['sal_max'].append(sal_max_byn)
            groups_byn[(spec, grd, current_geo)]['all_salaries'].append(sal_max_byn)

    results_byn = []
    for (spec, grd, current_geo), val_dict in groups_byn.items():
        min_list = val_dict['sal_min']
        max_list = val_dict['sal_max']
        all_list = val_dict['all_salaries']

        if min_list:
            min_median = median(min_list)
        else:
            min_median = None

        if max_list:
            max_median = median(max_list)
        else:
            max_median = None

        if all_list:
            market_median = median(all_list)
        else:
            market_median = None

        results_byn.append({
            'specialization': spec,
            'grade': grd,
            'geo': current_geo,
            'byn_min_median': int(round(min_median)) if min_median is not None else None,
            'byn_max_median': int(round(max_median)) if max_median is not None else None,
            'byn_market_median': int(round(market_median)) if market_median is not None else None,
        })

    # Получаем все уникальные валюты из результатов
    currencies = set()
    for row in results_gross:
        if 'currency' in row and row['currency']:
            currencies.add(row['currency'])
    
    # Получаем курсы только для используемых валют
    exchange_rates = ExchangeRate.objects.filter(currency__in=currencies).order_by('currency')
    
    context = {
        'results_gross': results_gross,
        'results_byn': results_byn,
        'exchange_rates': exchange_rates,
        'spec_filter': spec_filter,
        'grade_filter': grade_filter,
        'geo_filter': geo_filter,
    }
    return render(request, 'vacancies/summary.html', context)

#
# --- Остальные функции не изменены ---
#


def upload_excel(request):
    """
    Шаг 1: страница с формой ExcelUploadForm ->
    загружаем XLSX -> читаем -> чистим столбцы с формулами ->
    сохраняем в session -> редирект на preview_excel.
    """
    # (код остался без изменений)
    from apps.vacancies.forms import ExcelUploadForm

    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = form.cleaned_data['excel_file']

            wb = openpyxl.load_workbook(excel_file, data_only=False)
            sheet = wb.active

            raw_rows = []
            for row in sheet.iter_rows(values_only=False):
                row_cells = []
                for cell in row:
                    if cell.value is not None:
                        row_cells.append(str(cell.value).strip())
                    else:
                        row_cells.append("")
                if any(row_cells):
                    raw_rows.append(row_cells)

            max_cols = max(len(r) for r in raw_rows) if raw_rows else 0
            formula_columns = set()

            for row in raw_rows:
                for c_i, val in enumerate(row):
                    if c_i < max_cols and val.startswith("="):
                        formula_columns.add(c_i)

            cleaned_rows = []
            for row in raw_rows:
                new_row = []
                for c_i, val in enumerate(row):
                    if c_i not in formula_columns:
                        new_row.append(val)
                cleaned_rows.append(new_row)

            request.session['excel_rows'] = cleaned_rows
            return redirect('preview_excel')
        else:
            messages.error(request, "Форма ExcelUploadForm не валидна.")
    else:
        form = ExcelUploadForm()

    return render(request, 'vacancies/upload_excel.html', {'form': form})


def preview_excel(request):
    """
    Шаг 2: Показываем первую строку как заголовок и ТОЛЬКО 7 строк данных.
    Кнопка "Подтвердить" -> /process-excel/
    """
    rows_data = request.session.get('excel_rows')
    if not rows_data:
        messages.error(request, "Нет данных в Excel. Сначала загрузите файл.")
        return redirect('upload_excel')

    headers = rows_data[0] if rows_data else []
    data_rows = rows_data[1:8] if len(rows_data) > 1 else []

    return render(request, 'vacancies/preview_excel.html', {
        'headers': headers,
        'data_rows': data_rows
    })


def process_excel(request):
    """
    Шаг 3: Нажатие "Подтвердить" => прогоняем ВСЕ строки (кроме заголовков и пустых) через AI.
    """
    # (код остался без изменений)

    rows_data = request.session.get('excel_rows')
    if not rows_data:
        messages.error(request, "Нет данных для обработки.")
        return redirect('upload_excel')

    for row in rows_data[1:]:
        TaskQueue.create(data=" | ".join(row), priority=TaskQueue.Priority.MEDIUM)

    messages.success(request, f"Строки успешно добавлены в обработку.")
    request.session.pop('excel_rows', None)
    return redirect('index')


def handle():
    logger.info('Обновляем курсы валют')
    try:
        response = requests.get('https://api.nbrb.by/exrates/rates?periodicity=0', verify=False)
        if response.status_code == 200:
            data = response.json()
            updated_count = 1
            unsupported_currencies = set()
            
            # Добавляем BYN как базовую валюту
            ExchangeRate.objects.update_or_create(
                currency='BYN',
                defaults={'rate': 1.0}
            )
            
            for rate_data in data:
                try:
                    currency = rate_data.get('Cur_Abbreviation')
                    if not currency:
                        continue
                        
                    rate = rate_data.get('Cur_OfficialRate')
                    scale = rate_data.get('Cur_Scale', 1)
                    
                    if rate is None or scale == 0:
                        unsupported_currencies.add(currency)
                        continue
                    
                    rate = float(rate) / float(scale)
                    
                    ExchangeRate.objects.update_or_create(
                        currency=currency,
                        defaults={'rate': rate}
                    )
                    updated_count += 1
                    logger.info(f'Обновлен курс {currency}: {rate} BYN')
                except Exception as e:
                    logger.error(f'Ошибка при обработке валюты {currency}: {e}')
                    unsupported_currencies.add(currency)
            
            if unsupported_currencies:
                logger.warning(f'Не удалось обновить курсы для валют: {", ".join(unsupported_currencies)}')
            
            logger.info(f'Всего успешно обновлено {updated_count} курсов валют')
        else:
            logger.error(f'Ошибка API НБРБ: {response.status_code}')
    except Exception as e:
        logger.error(f"Ошибка при получении курсов валют: {e}")


def get_exchange_rates():
    """
    Получает курсы валют из базы данных.
    Возвращает словарь с курсами валют.
    Если курсы устарели (старше 24 часов), возвращает резервные значения.
    """
    cache_key = 'exchange_rates'
    rates = cache.get(cache_key)
    
    if rates is None:
        try:
            # Получаем все курсы из БД
            db_rates = ExchangeRate.objects.all()
            rates = {}
            
            # Проверяем актуальность курсов
            day_ago = timezone.now() - timedelta(days=1)
            
            for rate in db_rates:
                if rate.updated_at >= day_ago:
                    rates[rate.currency] = float(rate.rate)
            
            # Если есть актуальные курсы, кэшируем их на 1 час
            if rates:
                cache.set(cache_key, rates, timeout=3600)  # 1 час
            else:
                # Если курсы устарели, запускаем обновление и используем резервные значения
                handle()  # Пробуем обновить курсы
                rates = {'USD': 3.2, 'EUR': 3.5, 'RUB': 0.035, 'UZS': 0.00026, 'BYN': 1.0}
                
        except Exception as e:
            logger.error(f"Ошибка при получении курсов валют из БД: {e}")
            # В случае ошибки используем резервные значения
            rates = {'USD': 3.2, 'EUR': 3.5, 'RUB': 0.035, 'UZS': 0.00026, 'BYN': 1.0}
    
    return rates


@login_required
def profile(request):
    try:
        profile = request.user.profile
    except UserProfile.DoesNotExist:
        profile = UserProfile.objects.create(user=request.user)

    if request.method == 'POST':
        form = UserProfileForm(request.POST, instance=profile)
        if form.is_valid():
            profile = form.save()
            # Обновляем данные в модели User
            request.user.first_name = profile.first_name
            request.user.last_name = profile.last_name
            request.user.save()
            messages.success(request, 'Профиль успешно обновлен')
            return redirect('profile')
    else:
        # Инициализируем форму данными из профиля
        initial_data = {
            'last_name': profile.last_name or request.user.last_name,
            'first_name': profile.first_name or request.user.first_name,
            'middle_name': profile.middle_name,
            'company': profile.company,
            'phone': profile.phone,
        }
        form = UserProfileForm(instance=profile, initial=initial_data)

    return render(request, 'registration/profile.html', {'form': form})


@login_required
def edit_vacancy(request, vacancy_id):
    vacancy = get_object_or_404(Vacancy, pk=vacancy_id, is_active=True)
    if request.method == 'POST':
        form = VacancyEditForm(request.POST, instance=vacancy)
        if form.is_valid():
            vacancy = form.save(commit=False)
            vacancy.last_edited_by = request.user
            vacancy.save()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True})
            
            messages.success(request, 'Вакансия успешно обновлена')
            return redirect('vacancy_detail', vacancy_id=vacancy.id)
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                })
    else:
        form = VacancyEditForm(instance=vacancy)
    
    return render(request, 'vacancies/detail.html', {
        'form': form,
        'vacancy': vacancy,
        'edit_history': {
            'editor': vacancy.last_edited_by.get_full_name() if vacancy.last_edited_by else None,
            'edited_at': vacancy.last_edited_at
        }
    })
